# from flask import Flask, request, send_file
# from docxtpl import DocxTemplate
# import subprocess
# import os
# from dotenv import load_dotenv
# import base64

# app = Flask(__name__)

# @app.route('/pdf_generate', methods=['POST'])

# def pdf_generate():
    
#     # Base64-decode the file contents
#     template_file = request.form.get('file').encode('utf-8')
#     decode_template_file = base64.b64decode(template_file)
    
#     # Save the uploaded word file
#     with open(os.getenv('TEMPLATE_FILE'), 'wb') as f:
#         f.write(decode_template_file)
        
#     # Load the template
#     template = DocxTemplate(os.getenv('TEMPLATE_FILE'))
    
#     # Check if the template was loaded successfully
#     if template is None:
#         print('Failed to load the template file')
#         return send_file({'error': 'Failed to load the template file'})
    
#     # Define the data to be used for rendering the template
#     rendering_data = {}
    
#     # Extract all keys and values from the request.form and assign them to 
#     for key, value in request.form.items():
#         rendering_data[key] = value
        
#     # Render the template with the data
#     template.render(rendering_data)
#     template.save(os.getenv('RENDERED_WORD_FILE'))
    
#     command = ['libreoffice', '--convert-to', 'pdf:writer_pdf_Export', '--headless',
#                '--writer', '--nofirststartwizard', '--nolockcheck',
#                '--nologo', '--norestore', '--invisible',
#                '--outdir',  os.getenv('PDF_PATH'), os.getenv('RENDERED_WORD_FILE')]

#     subprocess.run(command, check=True)
    
#     # Check if the PDF file was successfully generated
#     if not os.path.exists(os.getenv('PDF_FILE')):
#         return send_file({'error': 'PDF file not generated'})
    
#     # return send_file(os.getenv('PDF_FILE'), as_attachment=True)
#     response = send_file(os.getenv('PDF_FILE'), as_attachment=True)
    
#     # Delete the PDF file after sending it
#     os.remove(os.getenv('PDF_FILE'))
     
#     return response

# if __name__ == '__main__':
#     app.run(host='0.0.0.0', port=5000)


##############################################################################################################


from flask import Flask, request, send_file, jsonify
from docxtpl import DocxTemplate
from openpyxl import Workbook
import subprocess
import os
from dotenv import load_dotenv
import base64

app = Flask(__name__)

@app.route('/generate_file', methods=['POST'])
def generate_file():
    data = request.get_json()
    
    if not data:
        return jsonify({'error': 'No JSON data provided'}), 400
    
    template_data = base64.b64decode(data.get('template', ''))
    output_format = data.get('output_format', '').lower()
    
    if not template_data or not output_format:
        return jsonify({'error': 'Missing template or output_format'}), 400
    
    template_path = os.path.join(os.getenv('UPLOAD_FOLDER'), 'template_file')
    with open(template_path, 'wb') as f:
        f.write(template_data)
    
    if output_format == 'pdf':
        generate_pdf(template_path)
        output_file = os.path.join(os.getenv('OUTPUT_FOLDER'), 'result.pdf')
    elif output_format == 'excel':
        generate_excel(template_path, data.get('rendering_data', {}))
        output_file = os.path.join(os.getenv('OUTPUT_FOLDER'), 'result.xlsx')
    elif output_format == 'word':
        generate_word(template_path, data.get('rendering_data', {}))
        output_file = os.path.join(os.getenv('OUTPUT_FOLDER'), 'result.docx')
    else:
        return jsonify({'error': 'Unsupported output_format'}), 400
    
    return send_file(output_file, as_attachment=True)

def generate_word(template_path, rendering_data):
    template = DocxTemplate(template_path)
    context = rendering_data
    
    # Render the template with the data
    template.render(context)
    output_path = os.path.join(os.getenv('OUTPUT_FOLDER'), 'result.docx')
    template.save(output_path)
    
def generate_excel(template_path, rendering_data):
    workbook = Workbook()
    sheet = workbook.active
    
    # Assign rendering_data to cells in the first column
    for idx, (key, value) in enumerate(rendering_data.items(), start=1):
        sheet.cell(row=idx, column=1, value=key)
        sheet.cell(row=idx, column=2, value=value)
    
    output_path = os.path.join(os.getenv('OUTPUT_FOLDER'), 'result.xlsx')
    workbook.save(output_path)

def generate_pdf(template_path):
    command = ['libreoffice', '--convert-to', 'pdf', '--headless', '--writer', '--outdir', os.getenv('OUTPUT_FOLDER'), template_path]
    subprocess.run(command, check=True)

if __name__ == '__main__':
    load_dotenv()
    app.run(host='0.0.0.0', port=5000)




